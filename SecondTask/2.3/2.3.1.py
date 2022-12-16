import csv
import re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
import matplotlib.pyplot as plt


def do_exit(message: str):
    """
    Выход из программы с выводом сообщения на консоль.
    :param message: Сообщение, выводимое на консоль.
    """
    print(message)
    exit()


class UserInput:
    """
    Класс для хранения данных, вводимых пользователем

    Attributes
    ----------
    file_name : str
        Путь до CSV-файла.
    profession : str
        Название профессии, введённое пользователем.
    """
    def __init__(self):
        self.file_name = input("Введите название файла: ")
        self.profession = input("Введите название профессии: ")


class Salary:
    """
    Класс для предоставления зарплаты.

    Attributes
    ----------
    salary_from : int
        Нижняя граница вилки оклада
    salary_to : int
        Верхняя граница вилки оклада
    salary_currency : str
        Валюта оклада
    average_salary: int
        Средняя зарплата, вычисляется по нижней и верхней границе
    currency_to_rub: dict
        Словарь, для перевода валюты в рубли
    """
    salary_from: str
    salary_to: str
    salary_currency: str
    average_salary: int
    currency_to_rub = {
        "AZN": 35.68,
        "BYR": 23.91,
        "EUR": 59.90,
        "GEL": 21.74,
        "KGS": 0.76,
        "KZT": 0.13,
        "RUR": 1,
        "UAH": 1.64,
        "USD": 60.66,
        "UZS": 0.0055,
    }

    def set_attribute(self, key: str, value: str):
        """
        Устанавливет значение поле зарплаты по ключу

        :param key: Название поля
        :param value: Значение поля
        """
        if key == "salary_gross":
            return
        if key != 'salary_currency':
            value = float(value)
        self.__setattr__(key, value)

    def set_average_salary(self):
        """
        Вычисляет значение средней зарплаты и устанавливает в
        аттрибут average_salary
        """
        self.average_salary = int(self.currency_to_rub[self.salary_currency] *
                                  (float(self.salary_from) + float(self.salary_to)) // 2)


class Vacancy:
    """
    Класс для предоставления вакансии

    Attributes
    ----------
    name: str
        Имя вакансии
    employer_name: str
        Имя компании
    salary: Salary
        Зарплата
    area_name: str
        Город
    published_at: int
        Год публикации
    date_time_publishing: datetime
        Дата публикации
    """
    name: str
    employer_name: str
    salary: Salary
    area_name: str
    published_at: int

    def __init__(self, fields: dict):
        """
        Инициализация класса с помощью словаря
        :param fields: Словарь с полями вакансии.
        Доступные ключи - name, salary_from, salary_to, salary_currency,
        area_name, published_at
        """
        self.date_time_publishing = None
        for key, value in fields.items():
            value = self.delete_html_tags(value)
            if not self.check_salary(key, value):
                self.__setattr__(key, value)

        self.published_time_formatter()
        self.salary.set_average_salary()

    def delete_html_tags(self, value: str) -> str:
        """
        Удаляет html теги
        :param value: строка с тегами
        :return: строка без тегов
        """
        value = re.sub('<.*?>', '', str(value)).replace("\r\n", "\n")
        value = ' '.join(value.split()).strip()
        return value

    def check_salary(self, key: str, value: str) -> bool:
        """
        Установка аттрибутов класса salary
        :param key: ключ словаря
        :param value: значение словаря
        :return: является ли ключ аттрибутом класса salary
        """
        if not key.__contains__("salary"):
            return False
        if not hasattr(self, "salary"):
            self.salary = Salary()
        self.salary.set_attribute(key, value)
        return True

    def published_time_formatter(self):
        """
        Установка аттрибутов date_time_publishing и
        published_at
        """
        hour, minute, second = self.published_at.split('T')[1].split('+')[0].split(':')
        year, month, day = self.published_at.split('T')[0].split('-')
        self.date_time_publishing = datetime(int(year), int(month), int(day), int(hour), int(minute), int(second))
        self.published_at = int(year)

    def get_field(self, field: str):
        """
        Получение аттрибута
        :param field: имя запрашиваемого аттрибута
        """
        if field == 'salary':
            return self.salary.average_salary
        return self.__getattribute__(field)


class DataDictionary:
    """
        Класс хранилища данных о вакансиях.

        Attributes
        ----------
        salary_years: dict
            Зарплаты по годам
        vacancies_years: dict
            Вакансии по годам
        salary_years_by_profession: dict
            Зарплаты по годам, по запрашиваемой профессии
        vacancies_years_by_profession: dict
            Вакансии по годам, по запрашиваемой профессии
        salaries_cities: dict
            Зарплаты по городам
        vacancy_cities_ratio: datetime
            Дата публикации
        """

    def __init__(self):
        """
        Инициализация аттрибутов
        """
        self.salary_years = {}  # a
        self.vacancies_years = {}  # b
        self.salary_years_by_profession = {}  # c
        self.vacancies_years_by_profession = {}  # d
        self.salaries_cities = {}  # e
        self.vacancy_cities_ratio = {}  # f
        self.city_vacancies_count = {}

    def update_data(self, vacancies: list, profession: str) -> None:
        """
        Инициализация словарей по списку вакансий
        :param vacancies: вакансии
        :param profession: название профессии
        """
        self.profession = profession
        for vacancy in vacancies:
            self.update_data_by_vacancy(vacancy, profession)

        self.correct_data(vacancies)

    def update_data_by_vacancy(self, vacancy, profession: str):
        """
        Инициализация словарей по вакансии
        :param vacancies: вакансии
        :param profession: название профессии
        """
        self.update_vacancies_count_dict('city_vacancies_count', 'area_name', vacancy)
        self.update_salary_dict('salary_years', 'published_at', vacancy)
        self.update_vacancies_count_dict('vacancies_years', 'published_at', vacancy)
        self.update_salary_dict('salaries_cities', 'area_name', vacancy)
        self.update_vacancies_count_dict('vacancy_cities_ratio', 'area_name', vacancy)
        if vacancy.name.__contains__(profession):
            self.update_salary_dict('salary_years_by_profession', 'published_at', vacancy)
            self.update_vacancies_count_dict('vacancies_years_by_profession', 'published_at', vacancy)

    def update_salary_dict(self, dict_name: str, field: str, vac: Vacancy) -> None:
        """
        Обновление словарей связанных с зарплатой
        :param dict_name: словарь
        :param field: поле, по которому формируется словарь
        :param vac: вакансия
        """
        dictionary = self.__getattribute__(dict_name)
        key = vac.get_field(field)
        if key not in dictionary.keys():
            dictionary[key] = [vac.salary.average_salary, 1]
        else:
            dictionary[key][0] += vac.salary.average_salary
            dictionary[key][1] += 1

    def update_vacancies_count_dict(self, dict_name: str, field: str, vac: Vacancy) -> None:
        """
        Обновление словарей связанных с подсчетом вакансий
        :param dict_name: словарь
        :param field: поле, по которому формируется словарь
        :param vac: вакансия
        """
        dictionary = self.__getattribute__(dict_name)
        key = vac.get_field(field)
        if key not in dictionary.keys():
            dictionary[key] = 1
        else:
            dictionary[key] += 1

    def correct_data(self, vacancies: list):
        """
        Корректирование словарей исходя из задачи
        :param vacancies: вакансии
        """
        for key, value in self.vacancy_cities_ratio.items():
            self.vacancy_cities_ratio[key] = round(value / len(vacancies), 4)

        buf = dict(sorted(self.salaries_cities.items(), key=lambda x: x[1][1] / x[1][0]))
        self.salaries_cities = self.get_first(buf, vacancies, 10)

        buf = dict(sorted(self.vacancy_cities_ratio.items(), key=lambda x: x[1], reverse=True))
        self.vacancy_cities_ratio = self.get_first(buf, vacancies, 10)

    def get_first(self, dictionary: dict, vacancies: list, amount: int) -> dict:
        """
        Получить первые amount полей
        :param dictionary: словарь
        :param vacancies: вакансии
        :param amount: количество вакансий
        """
        count = 0
        res = {}
        for key, value in dictionary.items():
            if count == amount:
                break
            if self.city_vacancies_count[key] >= len(vacancies) // 100:
                res[key] = value
                count += 1
        return res

    def print(self) -> None:
        """
        Вывод словарей
        """
        print_dictionary: {str, dict} = {
            "Динамика уровня зарплат по годам: ": self.salary_years,
            "Динамика количества вакансий по годам: ": self.vacancies_years,
            "Динамика уровня зарплат по годам для выбранной профессии: ": self.salary_years_by_profession,
            "Динамика количества вакансий по годам для выбранной профессии: ": self.vacancies_years_by_profession,
            "Уровень зарплат по городам (в порядке убывания): ": self.salaries_cities,
            "Доля вакансий по городам (в порядке убывания): ": self.vacancy_cities_ratio
        }
        for key, value in print_dictionary.items():
            if len(value) == 0:
                value = {k: 0 for k in self.salary_years.keys()}
            for k, v in value.items():
                if type(v) is list:
                    value[k] = v[0] // v[1]
            print(f"{key}{value}")


class DataSet:
    """
    Класс для считывания данных из файла
    """

    def csv_reader(file_name: str):
        """
        Считывания из csv файла
        :param file_name: имя файла
        :return: список вакансий
        """
        vacancies_array = []
        is_empty_file = True
        with open(file_name, 'r', newline='', encoding='utf-8-sig') as csvfile:
            reader = csv.DictReader(csvfile)
            if not reader.fieldnames:
                do_exit("Пустой файл")
            for row in reader:
                is_empty_file = False
                if all(row.values()):
                    vacancy = Vacancy(row)
                    vacancies_array.append(vacancy)
        if is_empty_file:
            do_exit("Нет данных")
        return vacancies_array


class Report:
    """
    Класс формирования отчёта по данным из DataDictionary

    Attributes
    ----------
    workbook : Workbook
        Класс, содержащий в себе функционал для работы с Excel таблицей.
    data : dict
        Словарь данных, получаемый из DataSet.
    """
    wb: Workbook

    def __init__(self, data: DataDictionary):
        """
        Инициализация
        :param data: словари из класса DataDictionary
        """
        self.work_book = Workbook()
        self.data = data

    def generate_excel(self):
        """
        Формирование excel документа
        """
        self.work_book.remove(self.work_book.active)
        self.generate_statistics_by_years()
        self.generate_statistics_by_cities()
        self.work_book.save("report.xlsx")

    def generate_statistics_by_years(self):
        """
        Формирование статистики по годам
        """
        ws = self.work_book.create_sheet("Статистика по годам")
        self.generate_data_dictionary(ws, "A", "Год", {v: str(k) for k, v in data.salary_years.items()})
        self.generate_data_dictionary(ws, "B", "Средняя зарплата", data.salary_years)
        self.generate_data_dictionary(ws, "C",
                                      f"Средняя зарплата - {data.profession}", data.salary_years_by_profession)
        self.generate_data_dictionary(ws, "D", "Количество вакансий", data.vacancies_years)
        self.generate_data_dictionary(ws, "E",
                                      f"Количество вакансий - {data.profession}", data.vacancies_years_by_profession)
        self.update_cell_settings(ws)

    def generate_data_dictionary(self, ws, column: str, name: str, dictionary: dict):
        """
        Формирование листа исходя из данных

        :param ws: лист
        :param column: столбец
        :param name: имя столбца
        :param dictionary: значения
        """
        ws[f"{column}1"] = name
        count = 2
        for year, value in dictionary.items():
            ws[f"{column}{count}"] = value
            count += 1

    def generate_statistics_by_cities(self):
        """
        Формирование статистики по городам
        """
        ws = self.work_book.create_sheet("Статистика по городам")
        self.generate_data_dictionary(ws, "A",
                                      "Город", {v: k for k, v in data.salaries_cities.items()})
        self.generate_data_dictionary(ws, "B", "Уровень зарплат", data.salaries_cities)
        self.generate_data_dictionary(ws, "D",
                                      "Город", {v: k for k, v in data.vacancy_cities_ratio.items()})
        self.generate_data_dictionary(ws, "E",
                                      "Доля вакансий", data.vacancy_cities_ratio)
        self.set_percent_style(ws)
        self.update_cell_settings(ws)

    def set_percent_style(self, ws):
        """
        Установка процентного стиля для листа ws
        :param ws: лист
        """
        for i in range(2, 12):
            ws[f"E{i}"].number_format = FORMAT_PERCENTAGE_00

    def update_cell_settings(self, ws):
        """
        Настройка ячеек
        :param ws: лист
        :return:
        """
        self.set_cell(ws)
        self.set_correctly_column_width(ws)

    def set_cell(self, ws):
        """
        Настройка параметров ячейки
        :param ws: лист
        :return:
        """
        isFirstCell = True
        for row in ws.rows:
            for cell in row:
                if cell.value:
                    cell.border = Border(top=Side(border_style="thin", color="000000"),
                                         left=Side(border_style="thin", color="000000"),
                                         right=Side(border_style="thin", color="000000"),
                                         bottom=Side(border_style="thin", color="000000"))
                    if isFirstCell:
                        cell.font = Font(bold=True)
            isFirstCell = False

    def set_correctly_column_width(self, ws):
        """
        Настройка ширины ячейки
        :param ws: лист
        :return:
        """
        a = {0: "A", 1: "B", 2: "C", 3: "D", 4: "E", 6: "F", 7: "G"}
        dims = {}
        for row in ws.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value)) + 1))

        for col, value in dims.items():
            ws.column_dimensions[a[col - 1]].width = value

    def generate_image(self):
        """
        Формирование изображения
        """
        fig, ax = plt.subplots(2, 2)
        self.generate_vertical_schedule(ax, data.salary_years, data.salary_years_by_profession,
                                        "Уровень зарплат по годам")
        self.generate_vertical_schedule(ax, data.vacancies_years, data.vacancies_years_by_profession,
                                        "Количество вакансий по годам")
        self.generate_horizontal_schedule(ax, "Уровень зарплат по городам")
        self.generate_circle_shedule(ax, "Доля вакансий по городам")
        plt.tight_layout()
        plt.savefig('graph.png')
        plt.show()

    def generate_vertical_schedule(self, ax, values, values_profession, title: str):
        """
        Сформировать вертикальную диаграмму
        :param ax: подобласть для отрисовки графика.
        :param values: значения
        :param values_profession: значения по профессии
        :param title: заголовок
        """
        my_label = ""
        coord = 0
        if title == "Уровень зарплат по годам":
            my_label = "Средняя зарплата"
        elif title == "Количество вакансий по годам":
            my_label = "Количество вакансий"
            coord = 1

        x_coord = [i for i in range(0, len(data.salary_years.keys()))]

        ax[0, coord].bar([x - 0.2 for x in x_coord], values.values(), width=0.5,
                         label=my_label)
        ax[0, coord].bar([x + 0.2 for x in x_coord], values_profession.values(), width=0.5,
                         label=f"{my_label} {data.profession}")
        ax[0, coord].set_xticks(x_coord, values.keys())
        ax[0, coord].set_xticklabels(values.keys(), rotation='vertical', va='top',
                                     ha='center')
        ax[0, coord].tick_params(axis='both', labelsize=8)
        ax[0, coord].legend(fontsize=8)
        ax[0, coord].grid(True, axis='y')
        ax[0, coord].set_title(title)

    def generate_horizontal_schedule(self, ax, title: str):
        """
        Сформировать горизонтальную диаграмму
        :param ax: подобласть для отрисовки графика.
        :param title: заголовок
        """
        ax[1, 0].invert_yaxis()
        ax[1, 0].tick_params(axis='both', labelsize=8)
        ax[1, 0].set_yticklabels(list(data.salaries_cities.keys()), fontsize=6, va='center', ha='right')
        ax[1, 0].barh(list(data.salaries_cities.keys()), list(data.salaries_cities.values()))
        ax[1, 0].grid(True, axis='x')
        ax[1, 0].set_title(title)

    def generate_circle_shedule(self, ax, title: str):
        """
        Сформировать круговую диаграмму
        :param ax: подобласть для отрисовки графика.
        :param title: заголовок
        """
        otherRatio = 1 - sum((list(data.vacancy_cities_ratio.values())))
        data.vacancy_cities_ratio.update({'Другие': otherRatio})
        ax[1, 1].pie(list(data.vacancy_cities_ratio.values()),
                     labels=list(data.vacancy_cities_ratio.keys()), textprops={'fontsize': 6})
        ax[1, 1].axis('scaled')
        ax[1, 1].set_title(title)


user_input = UserInput()
vacancies_array = DataSet.csv_reader(user_input.file_name)

if len(vacancies_array) == 0:
    do_exit("Ничего не найдено")

data = DataDictionary()
data.update_data(vacancies_array, user_input.profession)
data.print()

report = Report(data)
report.generate_image()
